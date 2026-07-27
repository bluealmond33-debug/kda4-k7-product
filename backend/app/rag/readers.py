"""포맷별 리더 — 원본 파일을 공통 중간형(Record)으로 읽는다.

지금 규정은 전부 PDF지만 부서에서 정리한 매뉴얼은 **xlsx·csv로 들어올 수 있다.**
그때 인제스트를 새로 짜지 않도록, 형식마다 이 파일에 리더 하나만 추가하면 되게 해 둔다.
구조화(조항/항목/내용/안내멘트)는 형식을 모르는 `structure.py`가 전담하므로,
리더가 할 일은 "행 혹은 블록을 Record로 만들어 주는 것"뿐이다.

의존성 원칙: **csv는 표준 라이브러리로 항상 되고, xlsx만 openpyxl을 지연 임포트**한다.
pdfplumber와 같은 방식이라, 라이브러리가 없는 호스트에서도 API는 그대로 뜬다.
"""

from __future__ import annotations

import csv
import io
from pathlib import Path
from typing import Any, Iterable

from .structure import Record, map_columns, normalize_header

# 헤더로 인정할 최소 매칭 수 — 한 칸만 우연히 맞은 행을 헤더로 오인하지 않게
_MIN_HEADER_HITS = 2
# 헤더를 찾기 위해 훑을 최대 행 수(엑셀은 제목·설명이 위에 몇 줄 붙는 경우가 흔하다)
_HEADER_SCAN_ROWS = 12


def _cells(row: Iterable[Any]) -> list[str]:
    return [("" if c is None else str(c)).strip() for c in row]


def find_header(rows: list[list[str]]) -> tuple[int, dict[str, int]]:
    """헤더 행 위치와 열 매핑을 찾는다.

    엑셀 매뉴얼은 첫 줄이 헤더가 아닌 경우가 많다(문서 제목·시트 설명이 먼저 온다).
    그래서 위에서부터 훑으며 **표준 열이 2개 이상 잡히는 첫 행**을 헤더로 본다.
    못 찾으면 (-1, {}) — 호출부가 '표가 아님'으로 처리한다.
    """
    for i, row in enumerate(rows[:_HEADER_SCAN_ROWS]):
        if not any(normalize_header(c) for c in row):
            continue
        cols = map_columns(row)
        if len(cols) >= _MIN_HEADER_HITS:
            return i, cols
    return -1, {}


def rows_to_records(rows: list[list[str]], section: str = "") -> list[Record]:
    """표(행 목록) → Record 목록. 헤더를 찾아 열을 표준 필드로 매핑한다."""
    hdr_idx, cols = find_header(rows)
    if hdr_idx < 0:
        return []

    out: list[Record] = []
    for r, row in enumerate(rows[hdr_idx + 1 :], start=hdr_idx + 2):
        vals = {k: (row[i] if i < len(row) else "") for k, i in cols.items()}
        # 내용·멘트가 모두 빈 행은 표의 여백이다(합계행·구분선 등) — 청크로 만들지 않는다
        if not (vals.get("content") or vals.get("script")):
            continue
        body = " ".join(v for v in (vals.get("content"), vals.get("script")) if v).strip()
        out.append(
            Record(
                body=body,
                page=1,
                section=vals.get("item") or section,
                kind="table",
                fields=vals,
                source_row=r,
            )
        )
    return out


def read_csv(path: str | Path, encoding: str | None = None) -> list[Record]:
    """CSV — 표준 라이브러리만 쓴다.

    한글 CSV는 UTF-8(BOM 포함)과 CP949가 섞여 들어온다. 인코딩을 지정하지 않으면
    순서대로 시도한다 — 여기서 실패하면 이후 단계가 통째로 깨지므로 관대하게 받는다.
    """
    raw = Path(path).read_bytes()
    encodings = [encoding] if encoding else ["utf-8-sig", "utf-8", "cp949", "euc-kr"]
    text = None
    for enc in encodings:
        try:
            text = raw.decode(enc)
            break
        except (UnicodeDecodeError, LookupError):
            continue
    if text is None:
        text = raw.decode("utf-8", errors="replace")

    rows = [_cells(r) for r in csv.reader(io.StringIO(text))]
    return rows_to_records(rows, section=Path(path).stem)


def read_xlsx(path: str | Path) -> list[Record]:
    """XLSX — 시트마다 표를 찾는다. 시트 이름이 곧 큰 항목(section)이 된다.

    openpyxl은 지연 임포트다(pdfplumber와 같은 규칙): 이 라이브러리가 없는 호스트에서도
    API는 정상 기동하고, 엑셀 인제스트를 시도할 때만 필요하다.
    """
    try:
        from openpyxl import load_workbook  # lazy
    except ImportError as e:  # pragma: no cover - 의존성 없는 호스트
        raise RuntimeError(
            "xlsx 인제스트에는 openpyxl이 필요합니다: pip install -r backend/requirements-rag.txt"
        ) from e

    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    out: list[Record] = []
    for ws in wb.worksheets:
        rows = [_cells(r) for r in ws.iter_rows(values_only=True)]
        recs = rows_to_records(rows, section=str(ws.title))
        # 시트가 여럿이면 시트 이름이 큰 항목이다 — 화면의 '큰 항목 > 작은 항목' 트리가 여기서 나온다
        for rec in recs:
            rec.fields.setdefault("sheet", str(ws.title))
        out.extend(recs)
    wb.close()
    return out


#: 확장자 → 리더. 새 형식은 여기 한 줄만 추가하면 된다.
READERS = {
    ".csv": read_csv,
    ".xlsx": read_xlsx,
    ".xlsm": read_xlsx,
}


def supported_suffixes() -> tuple[str, ...]:
    return tuple(READERS) + (".pdf",)


def read_tabular(path: str | Path) -> list[Record]:
    """표 형식(csv/xlsx) 파일을 Record로. PDF는 ingest.ingest_pdf가 맡는다."""
    suffix = Path(path).suffix.lower()
    reader = READERS.get(suffix)
    if not reader:
        raise ValueError(f"지원하지 않는 형식입니다: {suffix} (지원: {', '.join(supported_suffixes())})")
    return reader(path)
